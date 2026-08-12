#!/usr/bin/env python3
"""Turn the salon's DAILY SERVICE SALES REPORT workbooks into one paste-ready
SQL file for the Supabase SQL editor.

Usage:
    python3 xlsx_to_sql.py [-o import.sql] FILE.xlsx [FILE.xlsx ...]

Branch is read from the filename (MAIN_* / BRANCH_*). Reruns are safe: every
ticket carries a deterministic idempotency key and the SQL only inserts
tickets whose key is absent.

Source facts this transform is built around (verified against Jan+Feb 2026):
- Header order differs between months (SERVICE and TYPE OF SERVICE swap), so
  columns are mapped by header text per file.
- "DAILY SALES (2)" is a duplicate of "DAILY SALES" (same totals); only the
  latter is read.
- Most rows are anonymous daily tallies (no NAME). They become one ticket per
  row against a per-branch pooled walk-in client (clients.is_pool).
- Named rows group into one ticket per (date, name): the salon's history has
  no phone numbers, so identity is the normalised name.
- SHARING is a percent pair or "FIXED SHARING"; COMPANY SHARE is authoritative.
  The line's sharing_rate is fitted (6 dp) so the database's generated
  company_share_cents reproduces the workbook's peso split exactly.
- ASSIST is a peso deduction paid to the shampoo assistant out of the
  technician's share. The schema does not model intra-staff redistribution;
  totals are reported so nothing vanishes silently.
"""

import argparse
import hashlib
import re
import sys
from collections import Counter, defaultdict
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP

from openpyxl import load_workbook

TYPE_MAP = {"HAIR": "Hair", "NAILS": "Nail & Foot", "OTHERS": "Others"}

def q(s):
    """SQL string literal."""
    if s is None:
        return "null"
    return "'" + str(s).replace("'", "''") + "'"

def norm_name(s):
    return re.sub(r"\s+", " ", str(s).strip().upper())

def fit_rate(company_cents, total_cents):
    """Rate (6 dp) such that round-half-up(total*rate) == company, exactly."""
    if total_cents <= 0:
        return Decimal("0"), True
    base = (Decimal(company_cents) / Decimal(total_cents)).quantize(
        Decimal("0.000001"), rounding=ROUND_HALF_UP)
    step = Decimal("0.000001")
    for k in range(0, 40):
        for cand in ({base} if k == 0 else {base + k * step, base - k * step}):
            if cand < 0 or cand > 1:
                continue
            got = int((Decimal(total_cents) * cand).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
            if got == company_cents:
                return cand, True
    return base, False


def read_target(wb):
    """The month's TARGET figure from the DASHBOARD sheet (pesos), if present."""
    try:
        grid = [list(r) for r in wb["DASHBOARD"].iter_rows(values_only=True)]
    except KeyError:
        return None
    for i, row in enumerate(grid):
        for j, v in enumerate(row):
            if isinstance(v, str) and v.strip().upper() == "TARGET":
                below = grid[i + 3][j] if i + 3 < len(grid) and j < len(grid[i + 3]) else None
                if isinstance(below, (int, float)):
                    return int(round(below * 100))
    return None


def read_file(path, branch):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["DAILY SALES"]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip().upper() if h else "" for h in next(it)]

    def col(*names, optional=False):
        for n in names:
            if n in header:
                return header.index(n)
        if optional:
            return None
        raise SystemExit(f"{path}: none of {names} in header {header}")

    ix = {
        "date": col("DATE"), "name": col("NAME"), "brgy": col("BRGY."),
        "town": col("TOWN & PROVINCE"), "oldnew": col("OLD OR NEW"),
        "source": col("INQUIRY SOURCE"), "referred": col("REFERRED BY"),
        "service": col("SERVICE"), "type": col("TYPE OF SERVICE"),
        "amount": col("AMOUNT"), "no": col("NO.", "QTY"), "total": col("TOTAL AMOUNT"),
        "sharing": col("SHARING"), "company": col("COMPANY SHARE"),
        "tech": col("TECHNICIAN"),
        "assist": col("ASSIST", "ASSIST (BANLAW)", "MINUS ASSIST", optional=True),
        "rate": col("RATE"),
        # June 2026 onward: the salon upgraded its own tracking.
        "series": col("SERIES NO.", optional=True),
        "assisted_by": col("ASSITED BY", "ASSISTED BY", optional=True),
        "disc_type": col("TYPE OF DISCOUNT", optional=True),
        "disc_amount": col("DISCOUNT AMOUNT", optional=True),
        "online_method": col("ONLINE PAYMENT", optional=True),
        "online_amount": col("ONLINE PYMNT AMOUNT", "ONLINE PAYMENT AMOUNT", optional=True),
        # July 2026 onward: phones and service times arrive in the source.
        "phone": col("PHONE NUMBER", optional=True),
        "tstart": col("TIME STARTED (24-HR FORMAT)", "TIME STARTED", optional=True),
        "tend": col("TIME ENDED (24-HR FORMAT)", "TIME ENDED", optional=True),
    }

    def opt(r, key):
        i = ix[key]
        return r[i] if i is not None else None

    def phone_norm(v):
        """Excel drops leading zeros on numeric phones; recover 09XXXXXXXXX."""
        if v is None:
            return None
        digits = re.sub(r"\D", "", re.sub(r"\.0$", "", str(v).strip()))
        if len(digits) == 10 and digits.startswith("9"):
            digits = "0" + digits
        if len(digits) == 12 and digits.startswith("639"):
            digits = "0" + digits[2:]
        return digits if len(digits) == 11 and digits.startswith("09") else None

    def time_norm(v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.strftime("%H:%M")
        if hasattr(v, "hour") and hasattr(v, "minute"):
            return f"{v.hour:02d}:{v.minute:02d}"
        m = re.match(r"^(\d{1,2}):(\d{2})", str(v).strip())
        if m and 0 <= int(m.group(1)) <= 23:
            return f"{int(m.group(1)):02d}:{m.group(2)}"
        return None

    rows, skipped = [], []
    last_date = None
    for lineno, r in enumerate(it, start=2):
        total = r[ix["total"]]
        if not isinstance(total, (int, float)):
            continue
        d = r[ix["date"]]
        if isinstance(d, datetime):
            d = d.date()
        elif d is None:
            d = last_date
        if not isinstance(d, date):
            skipped.append((lineno, "no usable date"))
            continue
        last_date = d

        service = r[ix["service"]]
        tech = r[ix["tech"]]
        if not service or not tech:
            skipped.append((lineno, f"missing {'service' if not service else 'technician'}"))
            continue

        qty = r[ix["no"]]
        qty = int(qty) if isinstance(qty, (int, float)) and qty >= 1 else 1
        amount = r[ix["amount"]]
        company = r[ix["company"]]
        rating = r[ix["rate"]]
        try:
            rating = int(str(rating))
        except (TypeError, ValueError):
            rating = None
        if rating is not None and not (1 <= rating <= 5):
            rating = None

        rows.append({
            "lineno": lineno,
            "date": d,
            "name": str(r[ix["name"]]).strip() if r[ix["name"]] else None,
            "brgy": str(r[ix["brgy"]]).strip() if r[ix["brgy"]] else None,
            "town": str(r[ix["town"]]).strip() if r[ix["town"]] else None,
            "oldnew": (lambda v: v if v in ("OLD", "NEW") else None)(
                str(r[ix["oldnew"]]).strip().upper() if r[ix["oldnew"]] else None),
            "source": str(r[ix["source"]]).strip() if r[ix["source"]] else None,
            "referred": str(r[ix["referred"]]).strip() if r[ix["referred"]] else None,
            "service": norm_name(service),
            "type": TYPE_MAP.get(norm_name(r[ix["type"]]) if r[ix["type"]] else "", None),
            "sharing_label": str(r[ix["sharing"]]).strip() if r[ix["sharing"]] else None,
            "qty": qty,
            "amount_cents": int(round(float(amount) * 100)) if isinstance(amount, (int, float)) else None,
            "total_cents": int(round(float(total) * 100)),
            "company_cents": int(round(float(company) * 100)) if isinstance(company, (int, float)) else None,
            "tech": norm_name(tech),
            "assist_cents": int(round(float(opt(r, "assist")) * 100))
                if isinstance(opt(r, "assist"), (int, float)) else None,
            "rating": rating,
            "branch": branch,
            "series": (lambda v: re.sub(r"\.0$", "", str(v).strip()) if v is not None else None)(opt(r, "series")),
            "assisted_by": norm_name(opt(r, "assisted_by")) if opt(r, "assisted_by") else None,
            "disc_type": str(opt(r, "disc_type")).strip().upper() if opt(r, "disc_type") else None,
            "disc_amount_cents": int(round(float(opt(r, "disc_amount")) * 100))
                if isinstance(opt(r, "disc_amount"), (int, float)) else None,
            "online_method": str(opt(r, "online_method")).strip().upper() if opt(r, "online_method") else None,
            "online_amount_cents": int(round(float(opt(r, "online_amount")) * 100))
                if isinstance(opt(r, "online_amount"), (int, float)) else None,
            "phone": phone_norm(opt(r, "phone")),
            "phone_raw": str(opt(r, "phone")).strip() if opt(r, "phone") else None,
            "tstart": time_norm(opt(r, "tstart")),
            "tend": time_norm(opt(r, "tend")),
        })
    return rows, skipped, read_target(wb)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("files", nargs="+")
    ap.add_argument("-o", "--output", default="import_history.sql")
    args = ap.parse_args()

    all_rows, all_skipped = [], []
    targets = {}  # branch -> (latest month seen, target cents)
    for path in args.files:
        fname = path.rsplit("/", 1)[-1]
        m = re.search(r"(MAIN|BRANCH)_", fname)
        if not m:
            sys.exit(f"Cannot infer branch from filename: {fname}")
        rows, skipped, target = read_file(path, m.group(1))
        all_rows.extend(rows)
        all_skipped.extend((fname, ln, why) for ln, why in skipped)
        if rows and target:
            month = max(r["date"] for r in rows).strftime("%Y-%m")
            if m.group(1) not in targets or month > targets[m.group(1)][0]:
                targets[m.group(1)] = (month, target)
        print(f"{fname}: {len(rows)} rows, {len(skipped)} skipped", file=sys.stderr)

    # Fill missing type from the same service seen elsewhere; default Others.
    type_by_service = {}
    for r in all_rows:
        if r["type"]:
            type_by_service.setdefault(r["service"], Counter())[r["type"]] += 1
    for r in all_rows:
        if not r["type"]:
            c = type_by_service.get(r["service"])
            r["type"] = c.most_common(1)[0][0] if c else "Others"

    # Fit the sharing rate per line so generated company share is exact.
    unfit = 0
    for r in all_rows:
        company = r["company_cents"]
        if company is None:
            company = round(r["total_cents"] / 2)  # absent: assume 50-50, reported
            r["company_cents"] = company
            r["company_assumed"] = True
        rate, ok = fit_rate(company, r["total_cents"])
        r["rate6"] = f"{rate:.6f}"
        if not ok:
            unfit += 1

    # Ticket keys: the source's own series number is the strongest grouping
    # (June onward); before that, named rows group per (branch, date, name);
    # anonymous tallies are one ticket each.
    occurrence = Counter()
    for r in all_rows:
        if r["series"]:
            # Series numbers occasionally serve two clients (families on one
            # job order) or repeat across dates, so date and name are part of
            # the identity.
            who = hashlib.sha1(norm_name(r["name"]).encode()).hexdigest()[:12] if r["name"] else "anon"
            key = f"import:{r['branch']}:S:{r['series']}:{r['date']}:{who}"
        elif r["name"]:
            key = f"import:{r['branch']}:{r['date']}:N:{hashlib.sha1(norm_name(r['name']).encode()).hexdigest()[:16]}"
        else:
            sig = f"{r['branch']}|{r['date']}|{r['service']}|{r['tech']}|{r['qty']}|{r['total_cents']}"
            occurrence[sig] += 1
            key = f"import:{r['branch']}:{r['date']}:R:{hashlib.sha1(f'{sig}|{occurrence[sig]}'.encode()).hexdigest()[:16]}"
        r["ticket_key"] = key

    # ---------------------------------------------------------------- tickets
    # Ticket-level facts: discount allocation and the payment legs.
    DISC_MAP = {"STAFF DISC.": "staff", "SENIOR": "senior", "PWD": "pwd"}
    METHOD_MAP = {"GCASH": "gcash", "MAYA": "maya"}
    by_ticket = defaultdict(list)
    for r in all_rows:
        by_ticket[r["ticket_key"]].append(r)

    ticket_pay = {}      # key -> list of (method, cents)
    ticket_summary = {}  # key -> payment_method for the ticket header
    for key, rows in by_ticket.items():
        gross = sum(x["total_cents"] for x in rows)

        # Ticket discount: recorded once on one row of the ticket (verified
        # against DISCOUNT RATE x gross). Allocate pro-rata by line gross,
        # largest remainder, so the cents sum exactly.
        disc = sum(x["disc_amount_cents"] or 0 for x in rows)
        disc = min(disc, gross)
        if disc > 0 and gross > 0:
            shares = [(x["total_cents"] * disc) // gross for x in rows]
            rem = disc - sum(shares)
            order = sorted(range(len(rows)),
                           key=lambda i: (rows[i]["total_cents"] * disc) % gross,
                           reverse=True)
            for i in order[:rem]:
                shares[i] += 1
            dtype = next((x["disc_type"] for x in rows if x["disc_type"]), None)
            mapped = DISC_MAP.get(dtype or "", "promo")
            for x, extra in zip(rows, shares):
                x["alloc_disc_cents"] = extra
                x["alloc_disc_type"] = mapped if extra > 0 else None
        else:
            for x in rows:
                x["alloc_disc_cents"] = 0
                x["alloc_disc_type"] = None

        net = gross - disc

        # Online legs: usually repeated per line (sum == net), sometimes the
        # ticket total written once (max == net-ish). Clamp to net.
        cells = [x["online_amount_cents"] for x in rows if x["online_amount_cents"]]
        method = next((METHOD_MAP.get(x["online_method"] or "")
                       for x in rows if x["online_method"]), None)
        online = 0
        if cells and method:
            s, mx = sum(cells), max(cells)
            online = s if s <= net else (mx if mx <= net else net)
        cash = net - online
        legs = []
        if cash > 0:
            legs.append(("cash", cash))
        if online > 0:
            legs.append((method, online))
        if not legs:
            legs = [("cash", 0)]  # fully-discounted comp ticket
        ticket_pay[key] = legs
        ticket_summary[key] = ("split" if len(legs) > 1
                               else ("cash" if legs[0][0] == "cash" else "online"))

    # qty * unit - discount must equal the collected (net) amount exactly:
    # negotiated difference from the price column, plus the ticket discount
    # allocated to this line.
    for r in all_rows:
        qty, gross = r["qty"], r["total_cents"]
        unit = r["amount_cents"]
        if unit is None or unit * qty < gross:
            unit = -(-gross // qty)  # ceil division
        negotiated = unit * qty - gross
        r["unit_cents"] = unit
        r["net_cents"] = gross - r["alloc_disc_cents"]
        r["discount_cents"] = negotiated + r["alloc_disc_cents"]
        r["discount_type"] = (r["alloc_disc_type"]
                              if r["alloc_disc_cents"] > 0
                              else ("negotiated" if negotiated > 0 else None))

    # Per-service defaults: modal sharing rate, and one canonical type per
    # service name — the source files a service under different types on a
    # handful of rows (SHAVE under HAIR and NAILS), and a split identity
    # would fracture the analytics.
    svc_rates = defaultdict(Counter)
    svc_types_votes = defaultdict(Counter)
    for r in all_rows:
        svc_rates[r["service"]][r["rate6"]] += 1
        svc_types_votes[r["service"]][r["type"]] += 1
    svc_default_rate = {s: c.most_common(1)[0][0] for s, c in svc_rates.items()}
    svc_type = {s: c.most_common(1)[0][0] for s, c in svc_types_votes.items()}

    # Modal clean unit price per (branch, service) for the price list.
    price_votes = defaultdict(Counter)
    for r in all_rows:
        if r["discount_cents"] == 0 and r["amount_cents"] is not None:
            price_votes[(r["branch"], r["service"])][r["unit_cents"]] += 1
    modal_price = {k: c.most_common(1)[0][0] for k, c in price_votes.items()}

    # Validation expectations per branch-month. Revenue is NET of ticket
    # discounts (what was collected); company share is the fitted rate applied
    # to net, mirroring the database's generated column. The workbook's own
    # gross figures are carried alongside for the report.
    expect = defaultdict(lambda: {"revenue": 0, "company": 0, "treatments": 0, "rows": 0,
                                  "gross": 0, "discounts": 0, "company_gross": 0, "online": 0})
    assist_total = defaultdict(int)
    for r in all_rows:
        k = (r["branch"], r["date"].strftime("%Y-%m"))
        net = r["net_cents"]
        exp_company = int((Decimal(net) * Decimal(r["rate6"])).quantize(
            Decimal("1"), rounding=ROUND_HALF_UP))
        expect[k]["revenue"] += net
        expect[k]["company"] += exp_company
        expect[k]["gross"] += r["total_cents"]
        expect[k]["discounts"] += r["alloc_disc_cents"]
        expect[k]["company_gross"] += r["company_cents"]
        expect[k]["treatments"] += r["qty"]
        expect[k]["rows"] += 1
        if r["assist_cents"]:
            assist_total[k] += r["assist_cents"]
    for key, legs in ticket_pay.items():
        month = (by_ticket[key][0]["branch"], by_ticket[key][0]["date"].strftime("%Y-%m"))
        expect[month]["online"] += sum(c for m, c in legs if m != "cash")

    # ------------------------------------------------------------------ SQL
    out = []
    w = out.append
    w("-- inStyle Salon history import — generated by scripts/import/xlsx_to_sql.py")
    w(f"-- Files: {', '.join(p.rsplit('/', 1)[-1] for p in args.files)}")
    w("-- Rerun-safe: existing idempotency keys are skipped.")
    w("begin;")
    w("""
create temp table _imp (
  seq int, branch text, tdate date, ticket_key text,
  cname text, brgy text, town text, oldnew text, isource text, referred text,
  service text, stype text, qty int, unit_cents bigint, discount_cents bigint,
  discount_type text, rate6 numeric(8,6), tech text, assist_name text,
  rating int, net_cents bigint, series text, phone text, tstart text, tend text
) on commit drop;
""")

    CHUNK = 400
    for i in range(0, len(all_rows), CHUNK):
        vals = []
        for seq, r in enumerate(all_rows[i:i + CHUNK], start=i):
            vals.append(
                f"({seq},{q(r['branch'])},'{r['date']}',{q(r['ticket_key'])},"
                f"{q(r['name'])},{q(r['brgy'])},{q(r['town'])},{q(r['oldnew'])},"
                f"{q(r['source'])},{q(r['referred'])},{q(r['service'])},{q(r['type'])},"
                f"{r['qty']},{r['unit_cents']},{r['discount_cents']},{q(r['discount_type'])},"
                f"{r['rate6']},{q(r['tech'])},{q(r['assisted_by'])},"
                f"{r['rating'] if r['rating'] else 'null'},"
                f"{r['net_cents']},{q(r['series'])},{q(r['phone'])},{q(r['tstart'])},{q(r['tend'])})")
        w("insert into _imp values\n" + ",\n".join(vals) + ";")

    pay_vals = []
    for key, legs in sorted(ticket_pay.items()):
        for method, cents in legs:
            pay_vals.append(f"({q(key)},{q(method)},{cents},{q(ticket_summary[key])})")
    w("create temp table _payx (ticket_key text, method text, cents bigint, summary text) on commit drop;")
    for i in range(0, len(pay_vals), CHUNK):
        w("insert into _payx values\n" + ",\n".join(pay_vals[i:i + CHUNK]) + ";")

    w("""
-- Reference ids
create temp table _ctx on commit drop as
select
  (select id from businesses where code = 'SALON') as biz,
  (select id from profiles where role = 'owner' order by created_at limit 1) as owner;

do $$ begin
  if (select owner from _ctx) is null then
    raise exception 'No owner profile exists yet. Create the owner account first.';
  end if;
end $$;

-- Service types (Hair / Nail & Foot exist; Others may not)
insert into service_types (name, sort_order, business_id)
select v.stype, 50, (select biz from _ctx)
from (select distinct stype from _imp) v
where not exists (select 1 from service_types st
                  where st.name = v.stype and st.business_id = (select biz from _ctx));
""")

    svc_values = ",\n".join(
        f"({q(s)},{q(svc_type[s])},{svc_default_rate[s]})" for s in sorted(svc_type))
    w(f"""
-- Services observed in the history, with their modal company share
create temp table _svc (name text, stype text, rate numeric(8,6)) on commit drop;
insert into _svc values\n{svc_values};

insert into services (service_type_id, name, default_sharing_rate, default_duration_min)
select st.id, v.name, round(v.rate, 3), 45
from _svc v
join service_types st on st.name = v.stype and st.business_id = (select biz from _ctx)
on conflict (service_type_id, name) do nothing;
""")

    w("""
-- Technicians observed in the history (branch = where they mostly worked),
-- including assistants named on lines
insert into technicians (branch_id, full_name, active)
select b.id, v.tech, true
from (select distinct on (tech) tech, branch
      from (select tech, branch, count(*) c from (
              select tech, branch from _imp
              union all
              select assist_name, branch from _imp where assist_name is not null
            ) u group by 1, 2) x
      order by tech, c desc) v
join branches b on b.code = v.branch and b.business_id = (select biz from _ctx)
where not exists (
  select 1 from technicians t
  join branches tb on tb.id = t.branch_id
  where t.full_name = v.tech and tb.business_id = (select biz from _ctx)
);

-- The seeded placeholder roster and catalogue were stand-ins; retire any
-- that carry no history so pickers show the real salon.
update technicians t set active = false
from branches b
where b.id = t.branch_id and b.business_id = (select biz from _ctx)
  and t.full_name in ('Ana Ramos','Bea Salvador','Cristy Delos Reyes','Divine Ocampo',
                      'Elmer Padilla','Fely Baldoza','Grace Nolasco','Hazel Villar','Ivy Marquez')
  and not exists (select 1 from ticket_lines tl where tl.technician_id = t.id);

update services s set active = false
from service_types st
where st.id = s.service_type_id and st.business_id = (select biz from _ctx)
  and s.name not in (select name from _svc)
  and not exists (select 1 from ticket_lines tl where tl.service_id = s.id);

-- Pooled walk-in client per branch (money counts, retention excludes)
insert into clients (phone, phone_declined, full_name, is_pool)
select 'WALKIN-POOL-' || b.code, true, 'Walk-ins (unrecorded, ' || b.name || ')', true
from branches b
where b.business_id = (select biz from _ctx)
  and b.code in (select distinct branch from _imp)
on conflict (phone) do nothing;

-- Clients with a real phone number (July 2026 onward): phone is the
-- identity, exactly as the product's own rule says.
insert into clients (phone, phone_declined, full_name, barangay, town, inquiry_source,
                     first_visit_on, notes)
select distinct on (v.phone)
  v.phone, false, v.cname, v.brgy, v.town, v.isource, v.first_date,
  case when v.referred is not null then 'Referred by ' || v.referred else null end
from (
  select phone, cname, brgy, town, isource, referred,
         min(tdate) over (partition by phone) as first_date,
         seq
  from _imp where phone is not null
) v
order by v.phone, (v.cname is null), seq
on conflict (phone) do nothing;

-- Named clients without a phone: identity is the normalised name.
insert into clients (phone, phone_declined, full_name, barangay, town, inquiry_source,
                     first_visit_on, notes)
select
  'WALKIN-N-' || md5(upper(v.cname)),
  true,
  v.cname,
  v.brgy, v.town, v.isource,
  v.first_date,
  case when v.referred is not null then 'Referred by ' || v.referred else null end
from (
  select distinct on (upper(cname))
    cname,
    first_value(brgy) over w as brgy,
    first_value(town) over w as town,
    first_value(isource) over w as isource,
    first_value(referred) over w as referred,
    min(tdate) over (partition by upper(cname)) as first_date
  from _imp where cname is not null
  window w as (partition by upper(cname) order by (brgy is null), seq
               rows between unbounded preceding and unbounded following)
) v
on conflict (phone) do nothing;

-- Tickets: one per ticket_key not already imported
create temp table _new_tickets (id uuid, idempotency_key text) on commit drop;

with heads0 as (
  select distinct on (ticket_key)
    ticket_key, branch, tdate, cname, series,
    bool_or(oldnew = 'NEW') over (partition by ticket_key) as any_new,
    min(phone)  over (partition by ticket_key) as phone,
    min(tstart) over (partition by ticket_key) as tstart,
    max(tend)   over (partition by ticket_key) as tend
  from _imp
  order by ticket_key, seq
),
heads as (
  -- A paper series number occasionally covers two clients or two dates, and
  -- the counters can repeat across months, so duplicates — including tickets
  -- already imported by earlier batches — get a /n suffix.
  select h.*,
         case
           when series is null then null
           else series || case
             when row_number() over (partition by branch, series order by tdate, ticket_key)
                  + coalesce(prior.n, 0) = 1 then ''
             else '/' || (row_number() over (partition by branch, series order by tdate, ticket_key)
                          + coalesce(prior.n, 0))
           end
         end as series_unique
  from heads0 h
  left join lateral (
    select count(*) as n
    from tickets t
    join branches tb on tb.id = t.branch_id
    where tb.code = h.branch and tb.business_id = (select biz from _ctx)
      and (t.series_no = h.series or t.series_no like h.series || '/%')
  ) prior on h.series is not null
),
ins as (
  insert into tickets (branch_id, client_id, ticket_date, started_at, ended_at,
                       series_no, payment_method,
                       is_new_client, idempotency_key, created_by)
  select
    b.id,
    coalesce(ph.id, nc.id, pc.id),
    h.tdate,
    -- Service times arrive July onward; bad orderings are dropped, not fatal.
    case when h.tstart is not null
         then (h.tdate::text || ' ' || h.tstart || ':00+08')::timestamptz end,
    case when h.tstart is not null and h.tend is not null and h.tend >= h.tstart
         then (h.tdate::text || ' ' || h.tend || ':00+08')::timestamptz end,
    h.series_unique,
    coalesce((select px.summary from _payx px where px.ticket_key = h.ticket_key limit 1), 'cash'),
    coalesce(h.any_new, false) and h.cname is not null,
    h.ticket_key,
    (select owner from _ctx)
  from heads h
  join branches b on b.code = h.branch and b.business_id = (select biz from _ctx)
  left join clients ph on h.phone is not null and ph.phone = h.phone
  left join clients nc on h.phone is null and h.cname is not null
                      and nc.phone = 'WALKIN-N-' || md5(upper(h.cname))
  left join clients pc on h.cname is null and pc.phone = 'WALKIN-POOL-' || h.branch
  where not exists (select 1 from tickets t where t.idempotency_key = h.ticket_key)
  returning id, idempotency_key
)
insert into _new_tickets select id, idempotency_key from ins;

-- Lines for the tickets created in this run
insert into ticket_lines (ticket_id, service_id, technician_id, assist_technician_id,
                          qty, unit_price_cents,
                          discount_type, discount_cents, sharing_rate, rating, line_number)
select
  nt.id, s.id, tech.id,
  case when assist.id is distinct from tech.id then assist.id end,
  i.qty, i.unit_cents,
  i.discount_type,
  i.discount_cents, i.rate6, i.rating,
  row_number() over (partition by nt.id order by i.seq)
from _imp i
join _new_tickets nt on nt.idempotency_key = i.ticket_key
-- Lines resolve their service through the canonical mapping, not the row's
-- own type column, so a SHAVE filed under NAILS still lands on SHAVE.
join _svc sv on sv.name = i.service
join service_types st on st.name = sv.stype and st.business_id = (select biz from _ctx)
join services s on s.service_type_id = st.id and s.name = sv.name
join branches b on b.code = i.branch and b.business_id = (select biz from _ctx)
join lateral (
  select t.id from technicians t
  join branches tb on tb.id = t.branch_id
  where t.full_name = i.tech and tb.business_id = (select biz from _ctx)
  order by (t.branch_id = b.id) desc limit 1
) tech on true
left join lateral (
  select t.id from technicians t
  join branches tb on tb.id = t.branch_id
  where i.assist_name is not null and t.full_name = i.assist_name
    and tb.business_id = (select biz from _ctx)
  order by (t.branch_id = b.id) desc limit 1
) assist on true;

-- Assistants named in the source but not yet on the roster become
-- technicians first, so the join above can resolve them.

-- Payment legs (cash / gcash / maya) per new ticket
insert into ticket_payments (ticket_id, method, amount_cents)
select nt.id, px.method, px.cents
from _payx px join _new_tickets nt on nt.idempotency_key = px.ticket_key
where px.cents > 0
   or not exists (select 1 from _payx p2
                  where p2.ticket_key = px.ticket_key and p2.cents > 0);

-- Abort rather than commit a partial import: every staged row must have
-- landed, and every ticket's lines must equal its payment.
do $$
declare v_staged bigint; v_landed bigint; v_bad bigint;
begin
  select count(*) into v_staged from _imp i
  where exists (select 1 from _new_tickets nt where nt.idempotency_key = i.ticket_key);
  select count(*) into v_landed from ticket_lines tl
  where exists (select 1 from _new_tickets nt where nt.id = tl.ticket_id);
  if v_staged <> v_landed then
    raise exception 'Import aborted: % rows staged but % lines landed (service or technician failed to resolve).',
      v_staged, v_landed;
  end if;

  select count(*) into v_bad from _new_tickets nt
  where (select coalesce(sum(total_cents), 0) from ticket_lines where ticket_id = nt.id)
     <> (select coalesce(sum(amount_cents), 0) from ticket_payments where ticket_id = nt.id);
  if v_bad > 0 then
    raise exception 'Import aborted: % tickets have lines that do not equal their payment.', v_bad;
  end if;
end $$;
""")

    price_values = ",\n".join(
        f"({q(b)},{q(s)},{cents})" for (b, s), cents in sorted(modal_price.items()))
    w(f"""
-- Branch price list from the modal undiscounted charge per service
create temp table _prices (branch text, service text, cents bigint) on commit drop;
insert into _prices values\n{price_values};

insert into branch_service_prices (branch_id, service_id, price_cents, effective_from)
select b.id, s.id, p.cents, date '2026-01-01'
from _prices p
join branches b on b.code = p.branch and b.business_id = (select biz from _ctx)
join service_types st on st.business_id = (select biz from _ctx)
join services s on s.service_type_id = st.id and s.name = p.service
on conflict (branch_id, service_id, effective_from) do nothing;

""")
    for branch, (month, cents) in sorted(targets.items()):
        w(f"""-- Target from the {branch} {month} dashboard
update branches set monthly_target_cents = {cents}
where business_id = (select biz from _ctx) and code = {q(branch)};""")

    w("""
-- ---------------------------------------------------------------------------
-- Verification: these totals must match the workbook dashboards exactly.
-- ---------------------------------------------------------------------------
select b.code as branch,
       to_char(t.ticket_date, 'YYYY-MM') as month,
       sum(tl.total_cents) / 100.0 as sales,
       sum(tl.company_share_cents) / 100.0 as company_share,
       sum(tl.qty) as treatments,
       count(distinct t.id) as tickets
from tickets t
join branches b on b.id = t.branch_id
join ticket_lines tl on tl.ticket_id = t.id
where t.idempotency_key like 'import:%' and t.voided_at is null
group by 1, 2 order by 1, 2;

commit;""")

    with open(args.output, "w") as f:
        f.write("\n".join(out) + "\n")

    # ------------------------------------------------------------- report
    print(f"\nWrote {args.output} ({len(all_rows)} lines staged)", file=sys.stderr)
    print("\nExpected totals (compare with the verification query output):", file=sys.stderr)
    for (branch, month), e in sorted(expect.items()):
        print(f"  {branch:6s} {month}  sales(net) {e['revenue']/100:>12,.2f}  "
              f"company {e['company']/100:>12,.2f}  treatments {e['treatments']:>5}  "
              f"rows {e['rows']}", file=sys.stderr)
        if e['discounts'] or e['online']:
            print(f"         gross {e['gross']/100:>12,.2f}  discounts {e['discounts']/100:>10,.2f}  "
                  f"online {e['online']/100:>10,.2f}  company-if-gross {e['company_gross']/100:>12,.2f}",
                  file=sys.stderr)
    if unfit:
        print(f"\nWARNING: {unfit} lines could not fit an exact sharing rate", file=sys.stderr)
    assumed = sum(1 for r in all_rows if r.get("company_assumed"))
    if assumed:
        print(f"NOTE: {assumed} lines had no company share; assumed 50%", file=sys.stderr)
    with_phone = sum(1 for r in all_rows if r["phone"])
    bad_phone = [(r["phone_raw"]) for r in all_rows if r["phone_raw"] and not r["phone"]]
    if with_phone or bad_phone:
        print(f"\nPhones: {with_phone} rows with a valid number, {len(bad_phone)} unusable", file=sys.stderr)
        for s in list(dict.fromkeys(bad_phone))[:8]:
            print(f"  unusable: {s!r}", file=sys.stderr)
    timed = sum(1 for r in all_rows if r["tstart"])
    if timed:
        print(f"Service times present on {timed} rows", file=sys.stderr)
    if all_skipped:
        print(f"\nSkipped {len(all_skipped)} rows:", file=sys.stderr)
        for fname, ln, why in all_skipped[:20]:
            print(f"  {fname} line {ln}: {why}", file=sys.stderr)
    if assist_total:
        print("\nAssist (banlaw) deductions present in source but not modelled "
              "(intra-staff redistribution; company share unaffected):", file=sys.stderr)
        for (branch, month), cents in sorted(assist_total.items()):
            print(f"  {branch:6s} {month}  {cents/100:,.2f}", file=sys.stderr)


if __name__ == "__main__":
    main()

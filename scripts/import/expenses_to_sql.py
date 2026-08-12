#!/usr/bin/env python3
"""Daily Record expenses -> paste-ready SQL.

Reads the weekly "Daily Record / Daily Sales" blocks in the monthly workbooks
and turns the cash-deduction lines into `expenses` rows, so historical days
reconcile in the app the way they did on paper:

  Jan-Jun labels                Jul labels               category
  --------------                ----------               --------
  Withdrawals                   Withdrawal & Parcel      withdrawal
  Withdrawals from Bayarin      -                        withdrawal
  Allowance                     ALLOWANCE/SAHOD FD       allowance
  -                             SALON/SPA EXPENSES       supplies

Skipped on purpose: Discounts (already off the imported line totals), Gift
Certificates (a payment form, not a cash expense), and the per-technician
Bayarin/Cash Advance columns (intra-staff settlement, Stage 2).

The sheet's daily "Gcash Payments"/"Maya Payments" lines are what the drawer
was physically reconciled against, and the per-row online markers under-record
them (badly in Jan-May, slightly in Jun-Jul). So the SQL also trues up each
day's payment legs: cash legs are converted to gcash/maya (largest first, one
split at most) or back (smallest first) until the day's per-method totals
equal the sheet, then ticket headers are re-derived. Sales, lines, shares and
clients are untouched. The pass is convergent, so reruns are no-ops.

Every imported expense row's description ends with the marker "Daily Record
import"; the generated SQL deletes rows carrying the marker first, so reruns
are safe. The SQL stages each day's sheet "Net Total Cash" and finishes with
a per-month comparison of the app's expected cash against the sheet.

Usage: python3 expenses_to_sql.py <workbook.xlsx>... > import_expenses.sql
"""

import glob
import re
import sys
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP

import openpyxl

MARKER = "Daily Record import"

# label (lower-cased, prefix match) -> (category, description)
LINE_MAP = [
    ("withdrawals from bayarin", ("withdrawal", "Withdrawal from bayarin")),
    ("withdrawals",              ("withdrawal", "Withdrawals")),
    ("withdrawal & parcel",      ("withdrawal", "Withdrawal & parcel")),
    ("salon/spa expenses",       ("supplies",   "Salon/spa expenses")),
    ("allowance/sahod fd",       ("allowance",  "Allowance / sahod")),
    ("allowance",                ("allowance",  "Allowance")),
]

WEEK_SHEET = re.compile(r"^(Date - )?\d+-\d+")


def centavos(v) -> int:
    return int(Decimal(str(v)).scaleb(2).quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def block_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def scan_workbook(path: str, branch: str, month: int):
    """Yield (day, [(category, description, cents)...], sheet_net_cents)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    for sn in wb.sheetnames:
        if not WEEK_SHEET.match(sn):
            continue
        grid = [[c.value for c in row] for row in wb[sn].iter_rows()]
        for r, rowvals in enumerate(grid):
            for c, v in enumerate(rowvals):
                if not (isinstance(v, str) and "Daily Record" in v):
                    continue
                day = block_date(grid[r - 1][c]) if r >= 1 else None
                if day is None:
                    print(f"WARN {branch} {sn}: block without date at "
                          f"r{r + 1}c{c + 1}", file=sys.stderr)
                    continue
                # Trailing week sheets carry empty next-month template blocks.
                if (day.year, day.month) != (2026, month):
                    continue
                lines, net = [], None
                online = {"gcash payments": 0, "maya payments": 0}
                for rr in range(r + 1, min(r + 45, len(grid))):
                    lab = grid[rr][c] if c < len(grid[rr]) else None
                    if not isinstance(lab, str):
                        continue
                    lab_l = lab.strip().lower()
                    val = grid[rr][c + 2] if c + 2 < len(grid[rr]) else None
                    if lab_l == "net total cash" and net is None:
                        net = centavos(val) if isinstance(val, (int, float)) else 0
                    if lab_l in online and isinstance(val, (int, float)):
                        online[lab_l] = centavos(val)
                    for prefix, (cat, desc) in LINE_MAP:
                        if lab_l.startswith(prefix):
                            if isinstance(val, (int, float)) and val > 0:
                                lines.append((cat, desc, centavos(val)))
                            elif isinstance(val, (int, float)) and val < 0:
                                print(f"WARN {branch} {day}: negative "
                                      f"{lab.strip()} = {val}, skipped",
                                      file=sys.stderr)
                            break
                yield day, lines, net, (online["gcash payments"],
                                        online["maya payments"])
    wb.close()


def main(paths):
    files = []
    for p in paths:
        m = re.search(r"(MAIN|BRANCH)_DAILY.*?_(\d\d)\.2026", p.split("/")[-1])
        if not m:
            sys.exit(f"cannot tell branch/month from filename: {p}")
        files.append((m.group(1), int(m.group(2)), p))

    rows, days, online = [], {}, {}
    for branch, month, path in files:
        for day, lines, net, gm in scan_workbook(path, branch, month):
            key = (branch, day)
            if key in days:
                sys.exit(f"duplicate day block: {branch} {day}")
            days[key] = net
            online[key] = gm
            rows.extend((branch, day, cat, desc, cents)
                        for cat, desc, cents in lines)

    total = sum(r[4] for r in rows)
    print(f"-- Daily Record expenses: {len(rows)} rows over "
          f"{len(days)} day blocks, total {total / 100:,.2f}", file=sys.stderr)

    w = sys.stdout.write
    w("-- Historical Daily Record expenses (withdrawals, salon/spa expenses,\n")
    w("-- allowances) for Jan-Jul 2026, both branches. Generated by\n")
    w("-- scripts/import/expenses_to_sql.py; rerun-safe. Run as-is in the\n")
    w("-- Supabase SQL editor (default role, no impersonation).\n\n")

    w(f"delete from expenses where description like '%{MARKER}%';\n\n")

    w("drop table if exists _exp;\n")
    w("create temp table _exp (branch_code text, spent_on date,\n")
    w("  category text, description text, amount_cents bigint);\n")
    w("insert into _exp values\n")
    vals = [f"('{b}','{d}','{cat}','{desc} — {MARKER}',{cents})"
            for b, d, cat, desc, cents in rows]
    w(",\n".join(vals) + ";\n\n")

    w("drop table if exists _dr_days;\n")
    w("create temp table _dr_days (branch_code text, spent_on date,\n")
    w("  sheet_net_cents bigint, gcash_cents bigint, maya_cents bigint);\n")
    w("insert into _dr_days values\n")
    vals = [f"('{b}','{d}',{net if net is not None else 0},"
            f"{online[(b, d)][0]},{online[(b, d)][1]})"
            for (b, d), net in sorted(days.items())]
    w(",\n".join(vals) + ";\n\n")

    w("""-- True up each day's gcash/maya payment totals to the Daily Record.
do $$
declare
  t record; v_method text; v_delta bigint; leg record;
begin
  for t in
    select b.id as branch_id, d.branch_code, d.spent_on,
           d.gcash_cents, d.maya_cents
    from _dr_days d join branches b on b.code = d.branch_code
  loop
    foreach v_method in array array['gcash', 'maya'] loop
      v_delta := case v_method when 'gcash' then t.gcash_cents
                               else t.maya_cents end
               - coalesce((select sum(tp.amount_cents)
                           from ticket_payments tp
                           join tickets tk on tk.id = tp.ticket_id
                           where tk.branch_id = t.branch_id
                             and tk.ticket_date = t.spent_on
                             and tk.voided_at is null
                             and tp.method = v_method), 0);

      if v_delta > 0 then
        for leg in
          select tp.id, tp.amount_cents
          from ticket_payments tp
          join tickets tk on tk.id = tp.ticket_id
          where tk.branch_id = t.branch_id and tk.ticket_date = t.spent_on
            and tk.voided_at is null and tp.method = 'cash'
            and tp.amount_cents > 0
          order by tp.amount_cents desc, tp.id
        loop
          exit when v_delta = 0;
          if leg.amount_cents <= v_delta then
            update ticket_payments set method = v_method where id = leg.id;
            v_delta := v_delta - leg.amount_cents;
          else
            update ticket_payments
               set amount_cents = amount_cents - v_delta where id = leg.id;
            insert into ticket_payments (ticket_id, method, amount_cents)
            select ticket_id, v_method, v_delta
            from ticket_payments where id = leg.id;
            v_delta := 0;
          end if;
        end loop;
      elsif v_delta < 0 then
        for leg in
          select tp.id, tp.amount_cents
          from ticket_payments tp
          join tickets tk on tk.id = tp.ticket_id
          where tk.branch_id = t.branch_id and tk.ticket_date = t.spent_on
            and tk.voided_at is null and tp.method = v_method
          order by tp.amount_cents asc, tp.id
        loop
          exit when v_delta = 0;
          if leg.amount_cents <= -v_delta then
            update ticket_payments set method = 'cash' where id = leg.id;
            v_delta := v_delta + leg.amount_cents;
          else
            update ticket_payments
               set amount_cents = amount_cents + v_delta where id = leg.id;
            insert into ticket_payments (ticket_id, method, amount_cents)
            select ticket_id, 'cash', -v_delta
            from ticket_payments where id = leg.id;
            v_delta := 0;
          end if;
        end loop;
      end if;

      if v_delta <> 0 then
        raise exception 'online true-up unmet: % % % short by %',
          t.branch_code, t.spent_on, v_method, v_delta;
      end if;
    end loop;
  end loop;

  -- Merge duplicate same-method legs the conversions may have created,
  -- then re-derive ticket headers (cash / online / split only).
  with dup as (
    select tp.ticket_id, tp.method, min(tp.id::text)::uuid as keep_id,
           sum(tp.amount_cents) as total
    from ticket_payments tp
    join tickets tk on tk.id = tp.ticket_id
    join _dr_days d on d.spent_on = tk.ticket_date
    join branches b on b.id = tk.branch_id and b.code = d.branch_code
    group by tp.ticket_id, tp.method
    having count(*) > 1
  ), upd as (
    update ticket_payments tp set amount_cents = dup.total
    from dup where tp.id = dup.keep_id
    returning tp.id
  )
  delete from ticket_payments tp
  using dup
  where tp.ticket_id = dup.ticket_id and tp.method = dup.method
    and tp.id <> dup.keep_id;

  update tickets tk
     set payment_method = agg.derived
  from (
    select tp.ticket_id,
      case
        when bool_or(tp.method = 'cash' and tp.amount_cents > 0)
         and bool_or(tp.method in ('gcash','maya','bank','card')
                     and tp.amount_cents > 0) then 'split'
        when bool_or(tp.method in ('gcash','maya','bank','card')
                     and tp.amount_cents > 0) then 'online'
        else 'cash'
      end as derived
    from ticket_payments tp
    join tickets tk2 on tk2.id = tp.ticket_id
    join branches b on b.id = tk2.branch_id
    join _dr_days d on d.spent_on = tk2.ticket_date
                   and d.branch_code = b.code
    group by tp.ticket_id
    having bool_and(tp.method in ('cash','gcash','maya','bank','card'))
  ) agg
  where tk.id = agg.ticket_id
    and tk.payment_method in ('cash','online','split')
    and tk.payment_method is distinct from agg.derived;
end $$;

insert into expenses (branch_id, spent_on, category, amount_cents,
                      description, paid_from, recorded_by)
select b.id, e.spent_on, e.category, e.amount_cents, e.description, 'cash',
       (select id from profiles where role = 'owner' limit 1)
from _exp e
join branches b on b.code = e.branch_code;

do $$
declare v_staged bigint; v_landed bigint;
begin
  select count(*) into v_staged from _exp;
  select count(*) into v_landed from expenses
   where description like '%""" + MARKER + """%';
  if v_staged <> v_landed then
    raise exception 'staged % expense rows but % landed', v_staged, v_landed;
  end if;
end $$;

-- Verification: per month and branch, the imported totals and how many days
-- now reconcile exactly with the sheet's Net Total Cash. Days that differ
-- carry the workbook's own quirks (gift certificates, per-tech cash-advance
-- withholding) that the app deliberately does not model.
select
  to_char(d.spent_on, 'YYYY-MM')                as month,
  d.branch_code                                 as branch,
  count(*)                                      as days,
  count(*) filter (where v.expected_cash_cents = d.sheet_net_cents)
                                                as days_matching_sheet,
  coalesce(sum(e.amount_cents), 0) / 100.0      as expenses_imported,
  sum(abs(coalesce(v.expected_cash_cents, 0) - d.sheet_net_cents))
    filter (where v.expected_cash_cents is distinct from d.sheet_net_cents)
    / 100.0                                     as total_drift,
  count(*) filter (where coalesce(v.gcash_takings_cents, 0)
                       + coalesce(v.maya_takings_cents, 0)
                      <> d.gcash_cents + d.maya_cents)
                                                as online_off_days
from _dr_days d
join branches b on b.code = d.branch_code
left join v_daily_cash v
  on v.branch_id = b.id and v.business_date = d.spent_on
left join (select branch_code, spent_on, sum(amount_cents) as amount_cents
           from _exp group by 1, 2) e
  on e.branch_code = d.branch_code and e.spent_on = d.spent_on
group by 1, 2
order by 1, 2;
""")


if __name__ == "__main__":
    args = sys.argv[1:] or sorted(glob.glob("/root/.claude/uploads/*/*.xlsx"))
    main(args)
